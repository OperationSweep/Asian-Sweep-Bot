#!/usr/bin/env python3
"""Build the AMPS Kitting Report workbook (planned vs kitted, weekly + quarterly).

Inputs (paths passed as arguments):
  1. COOIS export  - Documented Goods Movements text export (.XLS tab-separated)
  2. ASP plan      - .xlsx with the weekly plan matrix

Usage: python3 build_excel_report.py <coois.xls> <asp.xlsx> <out.xlsx>

Business rules (validated July 2026 - see dashboard/README.md):
  - 1 work order = 1 AMP; a repeater needs FP-count AMPs.
  - Contract = last 3 digits of the 5-digit group in RRA/YAF material codes;
    GFF-variant -> contract mapping is learned from same-order co-occurrence.
  - Kitted = first net-positive GFF issue (261-262); Built = YAF 101 receipt.
  - Planned AMPs per week = ASP weekly repeater qty x that row's fibre pairs.
  - Quarter = ceil(ISO week / 13), capped at Q4, for plan and actuals alike.
"""
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime

import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

CONFIRMED_GFF = {"91CFL00302AFT": "429"}   # Aurora - confirmed by operations
KNOWN_NAMES = {"429": "Aurora", "295": "Anel Cam", "335": "Medusa", "588": "FIG"}

INK = "1F3864"
BLUE = "2A78D6"    # kitted series
GREEN = "008300"   # planned series
AR = Font(name="Arial", size=10)
ARB = Font(name="Arial", size=10, bold=True)
HDRF = Font(name="Arial", size=10, bold=True, color="FFFFFF")
HDRFILL = PatternFill("solid", fgColor=INK)
TITLE = Font(name="Arial", size=15, bold=True, color=INK)


def fam(m):
    return m[2:5]


def digits_of(m):
    g = re.search(r"\d{5}", m[5:])
    return g.group(0)[-3:] if g else None


def week_key(d):
    y, w, _ = d.isocalendar()
    return y, w


def quarter_of(week):
    return min(4, (week - 1) // 13 + 1)


def parse_coois(path):
    rows, header = [], None
    for ln in open(path, encoding="latin-1"):
        if "\t" not in ln:
            continue
        p = [x.strip() for x in ln.split("\t")]
        if p[1:2] == ["Order"]:
            header = p
            continue
        if header and len(p) >= len(header) - 1 and p[1].isdigit():
            rows.append(dict(zip(header, p)))
    for r in rows:
        r["date"] = datetime.strptime(r["Pstng Date"], "%d.%m.%Y").date()
    return rows


def attribute(rows):
    by_order = defaultdict(list)
    for r in rows:
        by_order[r["Order"]].append(r)
    votes = defaultdict(Counter)
    order_c = {}
    for o, rs in by_order.items():
        dg = {digits_of(r["Material"]) for r in rs if fam(r["Material"]) in ("RRA", "YAF")}
        dg.discard(None)
        if len(dg) == 1:
            c = dg.pop()
            order_c[o] = c
            for r in rs:
                if fam(r["Material"]) == "CFL":
                    votes[r["Material"]][c] += 1
    mapping = dict(CONFIRMED_GFF)
    for g, v in votes.items():
        mapping.setdefault(g, v.most_common(1)[0][0])
    orders = []
    for o, rs in by_order.items():
        contract = order_c.get(o)
        gffs = sorted({r["Material"] for r in rs if fam(r["Material"]) == "CFL"})
        if not contract:
            cands = {mapping.get(g) for g in gffs} - {None}
            contract = cands.pop() if len(cands) == 1 else None
        gff_net = sum(1 if r["Mvmt Type"] == "261" else -1
                      for r in rs if fam(r["Material"]) == "CFL" and r["Mvmt Type"] in ("261", "262"))
        issue_dates = sorted(r["date"] for r in rs if r["Mvmt Type"] == "261")
        kitted = issue_dates[0] if (gff_net > 0 or (gff_net == 0 and not any(
            fam(r["Material"]) == "CFL" for r in rs))) and issue_dates else None
        built_net = sum(1 if r["Mvmt Type"] == "101" else -1
                        for r in rs if fam(r["Material"]) == "YAF" and r["Mvmt Type"] in ("101", "102"))
        built_dates = sorted(r["date"] for r in rs
                             if fam(r["Material"]) == "YAF" and r["Mvmt Type"] == "101")
        built = built_dates[0] if built_net > 0 and built_dates else None
        orders.append({"order": o, "contract": contract, "kitted": kitted,
                       "built": built, "gff": "+".join(gffs)})
    return orders, mapping


def parse_asp(path):
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    grid = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i > 500:
            break
        grid.append(list(row))
    hi = next(i for i, r in enumerate(grid) if "PROJECT_ID" in [str(c) for c in r])
    hdr = [str(c) if c is not None else "" for c in grid[hi]]
    yr = grid[hi - 1]
    pq = hdr.index("Planned Qty")
    weekcols = []
    for j in range(pq + 1, len(hdr)):
        try:
            w, y = int(float(hdr[j])), int(float(yr[j]))
        except (TypeError, ValueError):
            continue
        if 1 <= w <= 53 and y >= 2000:
            weekcols.append((j, y, w))
    col = {n: hdr.index(n) for n in
           ("PROJECT_ID", "PROJECT_DESC", "ITEM_ID", "PROD_ID", "Total system", "Completed")}
    plan = []
    for r in grid[hi + 1:]:
        if not r[col["PROJECT_ID"]]:
            continue
        item = str(r[col["ITEM_ID"]] or "")
        fpm = re.search(r"[./](\d+)REP", item)
        prod = str(r[col["PROD_ID"]] or "")
        dg = re.search(r"\d{5}", prod)
        weekly = {}
        for j, y, w in weekcols:
            v = r[j] if j < len(r) else None
            if isinstance(v, (int, float)) and v > 0:
                weekly[(y, w)] = weekly.get((y, w), 0) + v
        plan.append({"name": str(r[col["PROJECT_DESC"]] or "").strip(),
                     "fp": int(fpm.group(1)) if fpm else None,
                     "digits": dg.group(0)[-3:] if dg else None,
                     "total": float(r[col["Total system"]] or 0),
                     "weekly": weekly})
    return plan


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HDRF
        cell.fill = HDRFILL


def build(coois_path, asp_path, out_path):
    rows = parse_coois(coois_path)
    orders, gff_map = attribute(rows)
    plan = parse_asp(asp_path)

    name_by_digits = dict(KNOWN_NAMES)
    for p in plan:
        if p["digits"] and p["name"]:
            name_by_digits.setdefault(p["digits"], p["name"])
    digits_by_name = {v.lower(): k for k, v in name_by_digits.items()}

    def contract_key(p):
        return p["digits"] or digits_by_name.get(p["name"].lower()) or "name:" + p["name"]

    def contract_label(key):
        if key.startswith("name:"):
            return key[5:]
        return name_by_digits.get(key, "Contract " + key)

    # weekly aggregation: (key, year, week) -> planned / kitted
    planned_w = defaultdict(float)
    plan_totals = defaultdict(lambda: {"reps": 0.0, "amps": 0.0})
    for p in plan:
        k = contract_key(p)
        plan_totals[k]["reps"] += p["total"]
        if p["fp"]:
            plan_totals[k]["amps"] += p["total"] * p["fp"]
            for (y, w), q in p["weekly"].items():
                planned_w[(k, y, w)] += q * p["fp"]
    kitted_w = defaultdict(int)
    built_w = defaultdict(int)
    act_totals = defaultdict(lambda: {"kitted": 0, "built": 0})
    for o in orders:
        k = o["contract"]
        if not k:
            continue
        if o["kitted"]:
            y, w = week_key(o["kitted"])
            kitted_w[(k, y, w)] += 1
            act_totals[k]["kitted"] += 1
        if o["built"]:
            y, w = week_key(o["built"])
            built_w[(k, y, w)] += 1
            act_totals[k]["built"] += 1

    keys = sorted(set(list(plan_totals) + list(act_totals)),
                  key=lambda k: -(act_totals[k]["kitted"] + plan_totals[k]["amps"]))
    all_weeks = sorted({(y, w) for (_, y, w) in list(planned_w) + list(kitted_w)})

    wb = openpyxl.Workbook()

    # ---------------- Weekly Data (values) ----------------
    wsd = wb.create_sheet("Weekly Data")
    hdr = ["Contract", "Year", "Week", "WeekKey", "Quarter", "Planned AMPs", "Kitted AMPs", "Built AMPs"]
    wsd.append(hdr)
    style_header(wsd, 1, len(hdr))
    for k in keys:
        for (y, w) in all_weeks:
            pv, kv, bv = planned_w.get((k, y, w), 0), kitted_w.get((k, y, w), 0), built_w.get((k, y, w), 0)
            if not (pv or kv or bv):
                continue
            wsd.append([contract_label(k), y, w, f"{y}-W{w:02d}", f"{y}-Q{quarter_of(w)}",
                        round(pv, 1), kv, bv])
    for row in wsd.iter_rows(min_row=2):
        for c in row:
            c.font = AR
    nW = wsd.max_row

    # ---------------- Dashboard ----------------
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_view.showGridLines = False
    ws["B2"] = "AMPS QUARTERLY KITTING DASHBOARD"
    ws["B2"].font = TITLE
    ws["B3"] = "Material Kitting Operations — Planned (ASP) vs Kitted (COOIS), per contract"
    ws["B3"].font = Font(name="Arial", size=10, color="52514E")
    snap = f"Data snapshot: COOIS {min(r['date'] for r in rows)} → {max(r['date'] for r in rows)}; " \
           f"ASP plan matrix from the pasted export. Regenerate with dashboard/build_excel_report.py."
    ws["B4"] = snap
    ws["B4"].font = Font(name="Arial", size=9, italic=True, color="898781")

    # KPI row (formulas over Weekly Data)
    kpis = [
        ("Total planned AMPs", f"=SUM('Weekly Data'!F2:F{nW})"),
        ("Total kitted AMPs", f"=SUM('Weekly Data'!G2:G{nW})"),
        ("Total built AMPs", f"=SUM('Weekly Data'!H2:H{nW})"),
        ("Overall % kitted vs plan", f"=IF(SUM('Weekly Data'!F2:F{nW})=0,\"\","
                                     f"SUM('Weekly Data'!G2:G{nW})/SUM('Weekly Data'!F2:F{nW}))"),
        ("Contracts tracked", len(keys)),
    ]
    r0 = 6
    for i, (label, val) in enumerate(kpis):
        c0 = 2 + i * 3
        cell = ws.cell(row=r0, column=c0, value=val)
        cell.font = Font(name="Arial", size=16, bold=True, color=INK)
        if "%" in label:
            cell.number_format = "0%"
        else:
            cell.number_format = "#,##0"
        lab = ws.cell(row=r0 + 1, column=c0, value=label)
        lab.font = Font(name="Arial", size=9, color="52514E")

    # Contract scoreboard
    r1 = r0 + 3
    ws.cell(row=r1, column=2, value="CONTRACT SCOREBOARD").font = ARB
    hdr2 = ["Contract", "Plan (reps)", "Plan (AMPs)", "Kitted (AMPs)", "Built (AMPs)",
            "% of plan", "Kitted this quarter", "Planned this quarter", "Verdict (cum. to date)"]
    for j, h in enumerate(hdr2):
        ws.cell(row=r1 + 1, column=2 + j, value=h)
    style_header(ws, r1 + 1, 1 + len(hdr2))
    ws.cell(row=r1 + 1, column=1).fill = PatternFill()   # keep col A unstyled
    today = datetime.utcnow().date()
    ty, tw = week_key(today)
    cur_q = f"{ty}-Q{quarter_of(tw)}"
    cur_wk = f"{ty}-W{tw:02d}"
    for i, k in enumerate(keys):
        rr = r1 + 2 + i
        label = contract_label(k)
        ws.cell(row=rr, column=2, value=label)
        ws.cell(row=rr, column=3, value=round(plan_totals[k]["reps"]) or None)
        ws.cell(row=rr, column=4,
                value=f"=SUMIFS('Weekly Data'!$F$2:$F${nW},'Weekly Data'!$A$2:$A${nW},$B{rr})")
        ws.cell(row=rr, column=5,
                value=f"=SUMIFS('Weekly Data'!$G$2:$G${nW},'Weekly Data'!$A$2:$A${nW},$B{rr})")
        ws.cell(row=rr, column=6,
                value=f"=SUMIFS('Weekly Data'!$H$2:$H${nW},'Weekly Data'!$A$2:$A${nW},$B{rr})")
        pa = plan_totals[k]["amps"]
        ws.cell(row=rr, column=7, value=f"=IF(D{rr}=0,\"\",E{rr}/D{rr})").number_format = "0%"
        ws.cell(row=rr, column=8,
                value=f"=SUMIFS('Weekly Data'!$G$2:$G${nW},'Weekly Data'!$A$2:$A${nW},$B{rr},"
                      f"'Weekly Data'!$E$2:$E${nW},\"{cur_q}\")")
        ws.cell(row=rr, column=9,
                value=f"=SUMIFS('Weekly Data'!$F$2:$F${nW},'Weekly Data'!$A$2:$A${nW},$B{rr},"
                      f"'Weekly Data'!$E$2:$E${nW},\"{cur_q}\")")
        # cumulative-to-date verdict: compare kitted vs planned for weeks <= current
        cum_p = f"SUMIFS('Weekly Data'!$F$2:$F${nW},'Weekly Data'!$A$2:$A${nW},$B{rr}," \
                f"'Weekly Data'!$D$2:$D${nW},\"<={cur_wk}\")"
        cum_k = f"SUMIFS('Weekly Data'!$G$2:$G${nW},'Weekly Data'!$A$2:$A${nW},$B{rr}," \
                f"'Weekly Data'!$D$2:$D${nW},\"<={cur_wk}\")"
        ws.cell(row=rr, column=10,
                value=f"=IF(AND({cum_p}=0,{cum_k}=0),\"not started\","
                      f"IF({cum_p}=0,\"AHEAD\",IF({cum_k}/{cum_p}<0.95,\"UNDER\","
                      f"IF({cum_k}/{cum_p}>1.05,\"AHEAD\",\"ON TRACK\"))))")
    last_score = r1 + 1 + len(keys)
    for row in ws.iter_rows(min_row=r1 + 2, max_row=last_score, min_col=2, max_col=10):
        for c in row:
            if c.font != ARB:
                c.font = AR
            if c.column in (3, 4, 5, 6, 8, 9):
                c.number_format = "#,##0"

    # ---------------- Weekly Totals (chart source, formulas) ----------------
    wst = wb.create_sheet("Weekly Totals")
    wst.append(["WeekKey", "Planned AMPs", "Kitted AMPs"])
    style_header(wst, 1, 3)
    for i, (y, w) in enumerate(all_weeks, start=2):
        wk = f"{y}-W{w:02d}"
        wst.cell(row=i, column=1, value=wk)
        wst.cell(row=i, column=2,
                 value=f"=SUMIFS('Weekly Data'!$F$2:$F${nW},'Weekly Data'!$D$2:$D${nW},A{i})")
        wst.cell(row=i, column=3,
                 value=f"=SUMIFS('Weekly Data'!$G$2:$G${nW},'Weekly Data'!$D$2:$D${nW},A{i})")
    for row in wst.iter_rows(min_row=2):
        for c in row:
            c.font = AR
    nT = wst.max_row

    # ---------------- Quarterly ----------------
    wsq = wb.create_sheet("Quarterly")
    wsq.append(["Quarter", "Planned AMPs", "Kitted AMPs", "% Kitted", "Built AMPs"])
    style_header(wsq, 1, 5)
    quarters = sorted({f"{y}-Q{quarter_of(w)}" for (y, w) in all_weeks})
    for i, q in enumerate(quarters, start=2):
        wsq.cell(row=i, column=1, value=q)
        wsq.cell(row=i, column=2,
                 value=f"=SUMIFS('Weekly Data'!$F$2:$F${nW},'Weekly Data'!$E$2:$E${nW},A{i})")
        wsq.cell(row=i, column=3,
                 value=f"=SUMIFS('Weekly Data'!$G$2:$G${nW},'Weekly Data'!$E$2:$E${nW},A{i})")
        wsq.cell(row=i, column=4, value=f"=IF(B{i}=0,\"\",MIN(1,C{i}/B{i}))").number_format = "0%"
        wsq.cell(row=i, column=5,
                 value=f"=SUMIFS('Weekly Data'!$H$2:$H${nW},'Weekly Data'!$E$2:$E${nW},A{i})")
    for row in wsq.iter_rows(min_row=2):
        for c in row:
            c.font = AR
    nQ = wsq.max_row

    # charts
    def styled_bar(title, ref_ws, min_col, max_col, min_row, max_row, cats_ref):
        ch = BarChart()
        ch.type = "col"
        ch.title = title
        ch.height, ch.width = 8, 24
        data = Reference(ref_ws, min_col=min_col, max_col=max_col, min_row=min_row - 1, max_row=max_row)
        ch.add_data(data, titles_from_data=True)
        ch.set_categories(cats_ref)
        ch.gapWidth = 60
        for s, color in zip(ch.series, (GREEN, BLUE)):
            s.graphicalProperties.solidFill = color
            s.graphicalProperties.line.noFill = True
        ch.y_axis.title = "AMPs"
        return ch

    ws.add_chart(styled_bar("Weekly — Planned vs Kitted AMPs", wst, 2, 3, 2, nT,
                            Reference(wst, min_col=1, min_row=2, max_row=nT)),
                 f"B{last_score + 2}")
    ws.add_chart(styled_bar("Quarterly — Planned vs Kitted AMPs", wsq, 2, 3, 2, nQ,
                            Reference(wsq, min_col=1, min_row=2, max_row=nQ)),
                 f"J{last_score + 2}")

    # ---------------- Orders (audit) ----------------
    wso = wb.create_sheet("Orders")
    wso.append(["Order", "Contract", "Kitted date", "Kitted week", "Built date", "GFF variant(s)"])
    style_header(wso, 1, 6)
    for o in sorted(orders, key=lambda x: x["order"]):
        wso.append([o["order"], contract_label(o["contract"]) if o["contract"] else "UNATTRIBUTED",
                    o["kitted"], f"{week_key(o['kitted'])[0]}-W{week_key(o['kitted'])[1]:02d}" if o["kitted"] else "",
                    o["built"], o["gff"]])
    for row in wso.iter_rows(min_row=2):
        for c in row:
            c.font = AR
            if c.column in (3, 5):
                c.number_format = "DD.MM.YYYY"

    # ---------------- Read Me ----------------
    wsr = wb.create_sheet("Read Me")
    notes = [
        ("AMPS KITTING REPORT — HOW TO READ IT", TITLE),
        ("", None),
        ("Dashboard: KPIs, contract scoreboard (formulas), weekly & quarterly planned-vs-kitted charts.", None),
        ("Weekly Data: one row per contract-week — the single source every formula aggregates from.", None),
        ("Orders: every work order with its contract, kitted week and GFF variant (audit trail).", None),
        ("", None),
        ("RULES", ARB),
        ("1 work order = 1 AMP; a repeater needs fibre-pair-count AMPs (24FP = 24 orders).", None),
        ("Planned AMPs/week = ASP weekly repeaters x that row's fibre pairs (from ITEM_ID).", None),
        ("Kitted = first net-positive GFF issue (261-262); Built = AMP assembly receipt (101).", None),
        ("Contract = last 3 digits of the 5-digit group in assembly codes; GFF variants are", None),
        ("contract-unique and the mapping is learned from the data (AFT->Aurora confirmed).", None),
        ("Quarter = ISO week 1-13 Q1, 14-26 Q2, 27-39 Q3, 40+ Q4.", None),
        ("Verdict: cumulative kitted vs cumulative plan to the current week, +/-5% = ON TRACK.", None),
        ("", None),
        ("KNOWN LIMITS", ARB),
        ("The ASP matrix only covers weeks present in the export - kitting before the matrix", None),
        ("window reads as AHEAD. The COOIS snapshot covers the export's date range only.", None),
        ("Contracts 438/517/521/527/627/746 have kitting but no matching ASP PROD_ID - names TBC.", None),
    ]
    for i, (txt, f) in enumerate(notes, start=1):
        wsr.cell(row=i, column=1, value=txt).font = f or AR
    wsr.column_dimensions["A"].width = 100

    # column widths
    for wsx, widths in ((ws, {2: 22, 3: 11, 4: 12, 5: 13, 6: 12, 7: 10, 8: 17, 9: 17, 10: 20}),
                        (wsd, {1: 22}), (wso, {1: 12, 2: 16, 3: 12, 4: 11, 5: 12, 6: 30}),
                        (wst, {1: 11, 2: 14, 3: 14}), (wsq, {1: 10, 2: 14, 3: 14, 4: 10, 5: 12})):
        for c, wdt in widths.items():
            wsx.column_dimensions[get_column_letter(c)].width = wdt

    wb.save(out_path)
    print(f"saved {out_path}: {len(orders)} orders, {len(keys)} contracts, "
          f"{nW - 1} weekly rows, {len(quarters)} quarters")


if __name__ == "__main__":
    build(sys.argv[1], sys.argv[2], sys.argv[3])
