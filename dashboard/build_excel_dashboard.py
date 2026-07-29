#!/usr/bin/env python3
"""Build the DYNAMIC AMPS Kitting Dashboard workbook (all live formulas).

Unlike build_excel_report.py (a static snapshot), this workbook is the
paste-and-refresh tool: every figure is an embedded Excel formula over the
'Paste COOIS' / 'Paste ASP' sheets, so pasting fresh exports recalculates
everything. Nothing is hardcoded except the editable Map sheet.

Usage:
  python3 build_excel_dashboard.py <coois.xls> <asp.xlsx> <out.xlsx>

Sheets:
  Start Here   - instructions
  Paste COOIS  - raw Documented Goods Movements export (preloaded)
  Paste ASP    - raw ASP plan grid (preloaded)
  Map          - GFF variant -> contract + fibre pair; contract digits -> name
  Engine       - one formula row per COOIS row (35,000 row capacity)
  Plan Engine  - one formula row per ASP row + numeric weekly-AMPs grid
  Dashboard    - KPIs, per-fibre-pair block, contract scoreboard,
                 weekly/quarterly tables + charts, week & quarter selectors
  Validation   - expected values computed from today's data for cross-check

Multi-fibre-pair rule: each FP design has its own GFF variant, so kitted
AMPs split per fibre pair via the Map (kitted AMPs = net GFF issues / 2).
A GFF row with FP left blank falls back to contract-level counting only.
"""
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime

import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.properties import CalcProperties

COOIS_CAP = 35000       # Engine formula capacity (rows of Paste COOIS)
ASP_CAP = 200           # Plan Engine capacity (rows of Paste ASP)
MAP_CAP = 40            # Map rows (GFF variants) incl. spares
WEEKS_SPAN = ("2024-W31", "2028-W52")

INK = "1F3864"
BLUE = "2A78D6"
GREEN = "008300"
AR = Font(name="Arial", size=10)
ARB = Font(name="Arial", size=10, bold=True)
HDRF = Font(name="Arial", size=10, bold=True, color="FFFFFF")
HDRFILL = PatternFill("solid", fgColor=INK)
TITLE = Font(name="Arial", size=15, bold=True, color=INK)
INPUT = Font(name="Arial", size=10, color="0000FF")
YELL = PatternFill("solid", fgColor="FFF2CC")

# seeded mapping: GFF variant -> (contract digits, contract name, FP or None)
SEED_GFF = [
    ("91CFL00302AFT", "429", "Aurora", 24),
    ("91CFL00302AFW", "588", "FIG", 24),
    ("91CFL00302AFM", "335", "Medusa", 19),
    ("91CFL00302AFY", "295", "Anel Cam", 6),
    ("91CFL00302AFZ", "295", "Anel Cam", 6),
    ("91CFL00302AFX", "438", "Contract 438", None),
    ("91CFL00302AFV", "627", "Contract 627", None),
    ("91CFL00302AED", "521", "Contract 521", None),
    ("91CFL00302AEM", "527", "Contract 527", None),
    ("91CFL00302AGB", "517", "Contract 517", None),
    ("91CFL00302AGA", "746", "Contract 746", None),
]


def week_list():
    out = []
    for y in range(2024, 2029):
        for w in range(1, 54):
            k = f"{y}-W{w:02d}"
            if WEEKS_SPAN[0] <= k <= WEEKS_SPAN[1]:
                out.append(k)
    return out


def quarter_list():
    return [f"{y}-Q{q}" for y in range(2025, 2029) for q in range(1, 5)]


def load_coois_grid(path):
    grid = []
    for ln in open(path, encoding="latin-1"):
        ln = ln.rstrip("\r\n")
        if "\t" not in ln:
            continue
        grid.append(ln.split("\t"))
    return grid


def load_asp_grid(path, max_rows=ASP_CAP):
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    grid = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= max_rows:
            break
        grid.append(list(row))
    return grid


def style_header(ws, row, ncols, start=1):
    for c in range(start, start + ncols):
        cell = ws.cell(row=row, column=c)
        cell.font = HDRF
        cell.fill = HDRFILL


def build(coois_path, asp_path, out_path):
    coois_grid = load_coois_grid(coois_path)
    asp_grid = load_asp_grid(asp_path)

    wb = openpyxl.Workbook()
    wb.calculation = CalcProperties(fullCalcOnLoad=True)

    # ---------------- Start Here ----------------
    st = wb.active
    st.title = "Start Here"
    lines = [
        ("AMPS QUARTERLY KITTING TRACKING DASHBOARD", TITLE),
        ("Store to Production Operations - Greenwich", ARB),
        ("", None),
        ("HOW TO USE - three steps", ARB),
        ("1. Paste your full COOIS 'Documented Goods Movements' export into 'Paste COOIS' cell A1", None),
        ("   (movements 261/262/101/102, all component materials, wide date range - all contracts at once).", None),
        ("2. Paste your full ASP plan into 'Paste ASP' cell A1 (the grid exactly as exported).", None),
        ("3. Press Ctrl+Alt+F9 to recalculate, then open the Dashboard tab.", None),
        ("", None),
        ("WHEN A NEW GFF VARIANT / CONTRACT APPEARS", ARB),
        ("Add one row on the 'Map' tab: GFF material code, contract digits, contract name, fibre pairs.", None),
        ("That is the ONLY maintenance this workbook ever needs.", None),
        ("", None),
        ("THE RULES BUILT IN (every cell is a live formula - click any number to trace it)", ARB),
        ("- 1 work order = 1 AMP and consumes exactly 2 GFFs, so kitted AMPs = net GFF issues (261-262) / 2.", None),
        ("- Each fibre-pair design has its own GFF variant -> kitted AMPs split per fibre pair via the Map.", None),
        ("- Planned AMPs per week = ASP weekly repeater qty x that row's fibre pairs (from ITEM_ID).", None),
        ("- Contract = last 3 digits of the 5-digit group (assembly codes and ASP PROD_ID agree).", None),
        ("- Quarter = ISO week 1-13 Q1, 14-26 Q2, 27-39 Q3, 40+ Q4.", None),
        ("- Capacity: %d COOIS rows, %d ASP rows. The Validation tab shows expected totals" % (COOIS_CAP, ASP_CAP), None),
        ("  for the data loaded on %s so you can confirm the formulas after any paste." % datetime.utcnow().date(), None),
    ]
    for i, (txt, f) in enumerate(lines, 1):
        st.cell(row=i, column=1, value=txt).font = f or AR
    st.column_dimensions["A"].width = 110

    # ---------------- Paste COOIS (preloaded raw) ----------------
    pc = wb.create_sheet("Paste COOIS")
    for i, row in enumerate(coois_grid, 1):
        for j, v in enumerate(row, 1):
            if v != "":
                pc.cell(row=i, column=j, value=v).font = AR
    n_coois = len(coois_grid)

    # ---------------- Paste ASP (preloaded raw) ----------------
    pa = wb.create_sheet("Paste ASP")
    for i, row in enumerate(asp_grid, 1):
        for j, v in enumerate(row, 1):
            if v is not None:
                pa.cell(row=i, column=j, value=v).font = AR
    # locate ASP layout for formulas
    hi = next(i for i, r in enumerate(asp_grid) if "PROJECT_ID" in [str(c) for c in r]) + 1  # 1-based header row
    hdr = [str(c) if c is not None else "" for c in asp_grid[hi - 1]]
    pq_col = hdr.index("Planned Qty") + 1          # 1-based
    item_col = hdr.index("ITEM_ID") + 1
    prod_col = hdr.index("PROD_ID") + 1
    desc_col = hdr.index("PROJECT_DESC") + 1
    total_col = hdr.index("Total system") + 1
    mat_first = pq_col + 1
    mat_last = len(hdr)
    n_mat = mat_last - mat_first + 1
    asp_first_data = hi + 1

    # ---------------- Map ----------------
    mp = wb.create_sheet("Map")
    mp.append(["GFF material code", "Contract digits", "Contract name", "Fibre pairs"])
    style_header(mp, 1, 4)
    for g, d, n, fp in SEED_GFF:
        mp.append([g, d, n, fp])
    for _ in range(MAP_CAP - len(SEED_GFF)):
        mp.append([None, None, None, None])
    for row in mp.iter_rows(min_row=2, max_row=MAP_CAP + 1):
        for c in row:
            c.font = INPUT
            c.fill = YELL
    mp.cell(row=MAP_CAP + 3, column=1,
            value="Add one row per GFF variant. Fibre pairs blank = counted at contract level only.").font = AR
    mp.cell(row=MAP_CAP + 5, column=1, value="CONTRACT LIST (drives the Dashboard scoreboard)").font = ARB
    mp.cell(row=MAP_CAP + 6, column=1, value="Contract name").font = HDRF
    mp.cell(row=MAP_CAP + 6, column=1).fill = HDRFILL
    contracts_start = MAP_CAP + 7
    seen = []
    for _, _, n, _ in SEED_GFF:
        if n not in seen:
            seen.append(n)
    extra = ["E2A", "I-AM", "Medusa_Ph2b", "TamTam", "WBA", "WKE", "Celia"]
    for n in extra:
        if n not in seen:
            seen.append(n)
    N_CONTRACTS = 30
    for i in range(N_CONTRACTS):
        c = mp.cell(row=contracts_start + i, column=1, value=seen[i] if i < len(seen) else None)
        c.font = INPUT
        c.fill = YELL
    for col, w in (("A", 20), ("B", 14), ("C", 18), ("D", 11)):
        mp.column_dimensions[col].width = w
    MAPR = f"Map!$A$2:$D${MAP_CAP + 1}"
    GFFCOL = f"Map!$A$2:$A${MAP_CAP + 1}"
    CONTRACTS = [f"Map!$A${contracts_start + i}" for i in range(N_CONTRACTS)]

    # ---------------- Engine ----------------
    en = wb.create_sheet("Engine")
    en.append(["GFF net", "Contract", "FP", "Date", "Week", "Quarter"])
    style_header(en, 1, 6)
    for r in range(2, COOIS_CAP + 2):
        s = r  # paste rows align 1:1 (header lines produce blanks, harmless)
        mat = f"'Paste COOIS'!$C{s}"
        mv = f"'Paste COOIS'!$I{s}"
        dt = f"'Paste COOIS'!$P{s}"
        en.cell(row=r, column=1,
                value=f'=IF(ISNA(MATCH({mat},{GFFCOL},0)),0,IF(({mv}&"")="261",1,IF(({mv}&"")="262",-1,0)))')
        en.cell(row=r, column=2, value=f'=IFERROR(VLOOKUP({mat},{MAPR},3,0),"")')
        en.cell(row=r, column=3, value=f'=IFERROR(VLOOKUP({mat},{MAPR},4,0),"")')
        en.cell(row=r, column=4,
                value=f'=IF({dt}="","",IF(ISNUMBER({dt}),{dt},'
                      f'DATE(VALUE(RIGHT({dt},4)),VALUE(MID({dt},4,2)),VALUE(LEFT({dt},2)))))')
        en.cell(row=r, column=5,
                value=f'=IF($D{r}="","",TEXT(YEAR($D{r}-WEEKDAY($D{r},2)+4),"0000")'
                      f'&"-W"&TEXT(ISOWEEKNUM($D{r}),"00"))')
        en.cell(row=r, column=6,
                value=f'=IF($E{r}="","",LEFT($E{r},4)&"-Q"&MIN(4,INT((VALUE(RIGHT($E{r},2))-1)/13)+1))')
    EN_NET = f"Engine!$A$2:$A${COOIS_CAP + 1}"
    EN_CON = f"Engine!$B$2:$B${COOIS_CAP + 1}"
    EN_FP = f"Engine!$C$2:$C${COOIS_CAP + 1}"
    EN_WK = f"Engine!$E$2:$E${COOIS_CAP + 1}"
    EN_Q = f"Engine!$F$2:$F${COOIS_CAP + 1}"

    # ---------------- Plan Engine ----------------
    pe = wb.create_sheet("Plan Engine")
    pe.cell(row=1, column=1, value="Weekly planned AMPs per ASP row (numeric mirror of the ASP matrix x fibre pairs)").font = ARB
    pe.cell(row=2, column=1, value="Contract").font = HDRF
    pe.cell(row=2, column=1).fill = HDRFILL
    pe.cell(row=2, column=2, value="FP").font = HDRF
    pe.cell(row=2, column=2).fill = HDRFILL
    pe.cell(row=2, column=3, value="Plan reps").font = HDRF
    pe.cell(row=2, column=3).fill = HDRFILL
    pe.cell(row=2, column=4, value="Plan AMPs").font = HDRF
    pe.cell(row=2, column=4).fill = HDRFILL
    # week-key header row for the matrix (row 2, cols 5..)
    for j in range(n_mat):
        pcol = get_column_letter(mat_first + j)
        cell = pe.cell(row=2, column=5 + j)
        cell.value = (f"=IF('Paste ASP'!{pcol}{hi}=\"\",\"\","
                      f"TEXT('Paste ASP'!{pcol}{hi - 1},\"0000\")&\"-W\""
                      f"&TEXT('Paste ASP'!{pcol}{hi},\"00\"))")
        cell.font = ARB
    pe_first = 3
    for i in range(ASP_CAP - asp_first_data + 1):
        r = pe_first + i
        s = asp_first_data + i
        prod = f"'Paste ASP'!${get_column_letter(prod_col)}{s}"
        desc = f"'Paste ASP'!${get_column_letter(desc_col)}{s}"
        item = f"'Paste ASP'!${get_column_letter(item_col)}{s}"
        tot = f"'Paste ASP'!${get_column_letter(total_col)}{s}"
        # contract name: via PROD_ID digits looked up in Map (cols B->C), else PROJECT_DESC
        pe.cell(row=r, column=1,
                value=f'=IF({desc}="","",IFERROR(VLOOKUP(RIGHT(MID({prod},6,5),3),'
                      f'Map!$B$2:$C${MAP_CAP + 1},2,0),{desc}))')
        pe.cell(row=r, column=2,
                value=f'=IFERROR(VALUE(TRIM(RIGHT(SUBSTITUTE(LEFT({item},FIND("REP",{item})-1),'
                      f'".",REPT(" ",10)),10))),"")')
        pe.cell(row=r, column=3, value=f'=IF({desc}="","",N({tot}))')
        pe.cell(row=r, column=4, value=f'=IF(OR($A{r}="",$B{r}=""),0,$C{r}*$B{r})')
        for j in range(n_mat):
            pcol = get_column_letter(mat_first + j)
            pe.cell(row=r, column=5 + j,
                    value=f'=IF(OR($A{r}="",$B{r}=""),0,IFERROR(N(\'Paste ASP\'!{pcol}{s})*$B{r},0))')
    pe_last = pe_first + (ASP_CAP - asp_first_data)
    PE_CON = f"'Plan Engine'!$A${pe_first}:$A${pe_last}"
    PE_FP = f"'Plan Engine'!$B${pe_first}:$B${pe_last}"
    PE_REPS = f"'Plan Engine'!$C${pe_first}:$C${pe_last}"
    PE_AMPS = f"'Plan Engine'!$D${pe_first}:$D${pe_last}"
    PE_KEYS = f"'Plan Engine'!$E$2:${get_column_letter(4 + n_mat)}$2"
    PE_GRID = f"'Plan Engine'!$E${pe_first}:${get_column_letter(4 + n_mat)}${pe_last}"

    # ---------------- Dashboard ----------------
    db = wb.create_sheet("Dashboard")
    db.sheet_view.showGridLines = False
    db["B2"] = "AMPS QUARTERLY KITTING TRACKING DASHBOARD"
    db["B2"].font = TITLE
    db["B3"] = "Planned (ASP) vs Kitted (COOIS) - every figure is a live formula"
    db["B3"].font = Font(name="Arial", size=10, color="52514E")
    db["B5"] = "SELECT WEEK:"
    db["B5"].font = ARB
    db["C5"] = "=Validation!B6"
    db["C5"].font = INPUT
    db["C5"].fill = YELL
    db["D5"] = "(type e.g. 2026-W29 - drives the 'Sel. week' columns)"
    db["D5"].font = Font(name="Arial", size=9, color="898781")
    db["B6"] = "SELECT QUARTER:"
    db["B6"].font = ARB
    db["C6"] = "=Validation!B7"
    db["C6"].font = INPUT
    db["C6"].fill = YELL
    db["D6"] = "(type e.g. 2026-Q3 - drives the 'Sel. quarter' columns)"
    db["D6"].font = Font(name="Arial", size=9, color="898781")
    SELW, SELQ = "Dashboard!$C$5", "Dashboard!$C$6"

    kpi_row = 8
    kpis = [
        ("Total kitted AMPs", f"=SUM({EN_NET})/2", "#,##0"),
        ("Total planned AMPs", f"=SUM({PE_AMPS})", "#,##0"),
        ("Overall % of plan", f"=IF(SUM({PE_AMPS})=0,\"\",SUM({EN_NET})/2/SUM({PE_AMPS}))", "0%"),
        ("Kitted in sel. week", f"=SUMIFS({EN_NET},{EN_WK},{SELW})/2", "#,##0"),
        ("Kitted in sel. quarter", f"=SUMIFS({EN_NET},{EN_Q},{SELQ})/2", "#,##0"),
    ]
    for i, (label, f, fmt) in enumerate(kpis):
        c0 = 2 + i * 2
        cell = db.cell(row=kpi_row, column=c0, value=f)
        cell.font = Font(name="Arial", size=16, bold=True, color=INK)
        cell.number_format = fmt
        db.cell(row=kpi_row + 1, column=c0, value=label).font = Font(name="Arial", size=9, color="52514E")

    # --- per fibre-pair block (one row per Map GFF row) ---
    fp_row = kpi_row + 3
    db.cell(row=fp_row, column=2, value="KITTING BY FIBRE-PAIR DESIGN (one row per GFF variant on the Map tab)").font = ARB
    fp_hdr = ["GFF variant", "Contract", "FP", "Kitted AMPs", "Kitted repeaters-worth",
              "Sel. week AMPs", "Sel. quarter AMPs"]
    for j, h in enumerate(fp_hdr):
        db.cell(row=fp_row + 1, column=2 + j, value=h)
    style_header(db, fp_row + 1, len(fp_hdr), start=2)
    PCM = f"'Paste COOIS'!$C$2:$C${COOIS_CAP + 1}"
    PCI = f"'Paste COOIS'!$I$2:$I${COOIS_CAP + 1}"
    for i in range(MAP_CAP):
        r = fp_row + 2 + i
        m = i + 2
        db.cell(row=r, column=2, value=f'=IF(Map!A{m}="","",Map!A{m})')
        db.cell(row=r, column=3, value=f'=IF(Map!A{m}="","",Map!C{m})')
        db.cell(row=r, column=4, value=f'=IF(Map!A{m}="","",Map!D{m})')
        net_expr = (f'SUMPRODUCT(({PCM}=Map!A{m})*'
                    f'((({PCI}&"")="261")-(({PCI}&"")="262")))')
        db.cell(row=r, column=5, value=f'=IF(Map!A{m}="","",{net_expr}/2)')
        db.cell(row=r, column=6, value=f'=IF(OR(Map!A{m}="",Map!D{m}=""),"",E{r}/Map!D{m})')
        db.cell(row=r, column=7,
                value=f'=IF(Map!A{m}="","",SUMPRODUCT(({PCM}=Map!A{m})*'
                      f'((({PCI}&"")="261")-(({PCI}&"")="262"))*({EN_WK}={SELW}))/2)')
        db.cell(row=r, column=8,
                value=f'=IF(Map!A{m}="","",SUMPRODUCT(({PCM}=Map!A{m})*'
                      f'((({PCI}&"")="261")-(({PCI}&"")="262"))*({EN_Q}={SELQ}))/2)')
    fp_last = fp_row + 1 + MAP_CAP

    # --- contract scoreboard ---
    sc_row = fp_last + 2
    db.cell(row=sc_row, column=2, value="CONTRACT SCOREBOARD (rows driven by the Contract list on the Map tab)").font = ARB
    sc_hdr = ["Contract", "Plan (reps)", "Plan (AMPs)", "Kitted (AMPs)", "% of plan",
              "Sel. week planned", "Sel. week kitted", "Sel. quarter planned", "Sel. quarter kitted",
              "Verdict (cumulative to sel. week)"]
    for j, h in enumerate(sc_hdr):
        db.cell(row=sc_row + 1, column=2 + j, value=h)
    style_header(db, sc_row + 1, len(sc_hdr), start=2)
    for i in range(N_CONTRACTS):
        r = sc_row + 2 + i
        cref = CONTRACTS[i]
        db.cell(row=r, column=2, value=f"=IF({cref}=\"\",\"\",{cref})")
        db.cell(row=r, column=3, value=f"=IF({cref}=\"\",\"\",SUMIFS({PE_REPS},{PE_CON},{cref}))")
        db.cell(row=r, column=4, value=f"=IF({cref}=\"\",\"\",SUMIFS({PE_AMPS},{PE_CON},{cref}))")
        db.cell(row=r, column=5, value=f"=IF({cref}=\"\",\"\",SUMIFS({EN_NET},{EN_CON},{cref})/2)")
        db.cell(row=r, column=6, value=f"=IF(OR({cref}=\"\",D{r}=0),\"\",E{r}/D{r})").number_format = "0%"
        db.cell(row=r, column=7,
                value=f"=IF({cref}=\"\",\"\",SUMPRODUCT(({PE_CON}={cref})*IFERROR(INDEX({PE_GRID},0,"
                      f"MATCH({SELW},{PE_KEYS},0)),0)))")
        db.cell(row=r, column=8, value=f"=IF({cref}=\"\",\"\",SUMIFS({EN_NET},{EN_CON},{cref},{EN_WK},{SELW})/2)")
        db.cell(row=r, column=9,
                value=f"=IF({cref}=\"\",\"\",SUMPRODUCT(({PE_CON}={cref})*({PE_GRID})*"
                      f"(LEFT({PE_KEYS},4)&\"-Q\"&(INT((VALUE(RIGHT({PE_KEYS},2))-1)/13)+1-(VALUE(RIGHT({PE_KEYS},2))>52))={SELQ})))")
        db.cell(row=r, column=10, value=f"=IF({cref}=\"\",\"\",SUMIFS({EN_NET},{EN_CON},{cref},{EN_Q},{SELQ})/2)")
        cum_p = f"SUMPRODUCT(({PE_CON}={cref})*({PE_GRID})*({PE_KEYS}<={SELW}))"
        cum_k = f"SUMPRODUCT(({EN_CON}={cref})*({EN_NET})*({EN_WK}<={SELW})*({EN_WK}<>\"\"))/2"
        db.cell(row=r, column=11,
                value=f"=IF({cref}=\"\",\"\",IF(AND({cum_p}=0,{cum_k}=0),\"not started\","
                      f"IF({cum_p}=0,\"AHEAD\",IF({cum_k}/{cum_p}<0.95,\"UNDER\","
                      f"IF({cum_k}/{cum_p}>1.05,\"AHEAD\",\"ON TRACK\")))))")
    sc_last = sc_row + 1 + N_CONTRACTS

    # --- weekly table + chart ---
    wt = wb.create_sheet("Weekly")
    wt.append(["Week", "Planned AMPs", "Kitted AMPs"])
    style_header(wt, 1, 3)
    wks = week_list()
    for i, k in enumerate(wks, start=2):
        wt.cell(row=i, column=1, value=k)
        wt.cell(row=i, column=2,
                value=f"=IFERROR(SUM(INDEX({PE_GRID},0,MATCH(A{i},{PE_KEYS},0))),0)")
        wt.cell(row=i, column=3, value=f"=SUMIFS({EN_NET},{EN_WK},A{i})/2")
        for c in wt[i]:
            c.font = AR
    nWT = len(wks) + 1

    qt = wb.create_sheet("Quarterly Tbl")
    qt.append(["Quarter", "Planned AMPs", "Kitted AMPs", "% Kitted"])
    style_header(qt, 1, 4)
    qs = quarter_list()
    for i, q in enumerate(qs, start=2):
        qt.cell(row=i, column=1, value=q)
        qt.cell(row=i, column=2,
                value=f"=SUMPRODUCT(({PE_GRID})*(LEFT({PE_KEYS},4)&\"-Q\"&"
                      f"(INT((VALUE(RIGHT({PE_KEYS},2))-1)/13)+1-(VALUE(RIGHT({PE_KEYS},2))>52))=A{i}))")
        qt.cell(row=i, column=3, value=f"=SUMIFS({EN_NET},{EN_Q},A{i})/2")
        qt.cell(row=i, column=4, value=f"=IF(B{i}=0,\"\",MIN(1,C{i}/B{i}))").number_format = "0%"
        for c in qt[i]:
            c.font = AR
    nQT = len(qs) + 1

    def styled_bar(title, src, ncols, nrows):
        ch = BarChart()
        ch.type = "col"
        ch.title = title
        ch.height, ch.width = 8, 24
        data = Reference(src, min_col=2, max_col=1 + ncols, min_row=1, max_row=nrows)
        ch.add_data(data, titles_from_data=True)
        ch.set_categories(Reference(src, min_col=1, min_row=2, max_row=nrows))
        ch.gapWidth = 60
        for s, color in zip(ch.series, (GREEN, BLUE)):
            s.graphicalProperties.solidFill = color
            s.graphicalProperties.line.noFill = True
        ch.y_axis.title = "AMPs"
        return ch

    db.add_chart(styled_bar("Weekly - Planned vs Kitted AMPs", wt, 2, nWT), f"B{sc_last + 2}")
    db.add_chart(styled_bar("Quarterly - Planned vs Kitted AMPs", qt, 2, nQT), f"J{sc_last + 2}")

    # formatting for dashboard numbers
    for row in db.iter_rows(min_row=fp_row + 2, max_row=sc_last, min_col=2, max_col=11):
        for c in row:
            if c.font != ARB and c.fill.fgColor.rgb != "00FFF2CC":
                c.font = AR
            if c.column in (3, 4, 5, 7, 8, 9, 10):
                c.number_format = "#,##0"
    for col, w in (("B", 20), ("C", 15), ("D", 14), ("E", 13), ("F", 18), ("G", 15),
                   ("H", 17), ("I", 17), ("J", 17), ("K", 24)):
        db.column_dimensions[col].width = w

    # ---------------- Validation (expected values for today's data) ----------------
    # compute expectations in python with the same rules
    exp_orders = Counter()
    gff_map = {g: (n, fp) for g, d, n, fp in SEED_GFF}
    net_by_c = Counter()
    for row in coois_grid:
        if len(row) < 16 or not row[1].strip().isdigit():
            continue
        mat, mv = row[2].strip(), row[8].strip()
        if mat in gff_map and mv in ("261", "262"):
            net_by_c[gff_map[mat][0]] += 1 if mv == "261" else -1
    va = wb.create_sheet("Validation")
    va.append(["Check", "Expected (computed %s)" % datetime.utcnow().date(), "Live formula", "Match?"])
    style_header(va, 1, 4)
    total_amps = sum(net_by_c.values()) / 2
    va.append(["Total kitted AMPs (all contracts)", total_amps,
               f"=SUM({EN_NET})/2", "=IF(B2=C2,\"OK\",\"CHECK\")"])
    r = 3
    for cname, net in sorted(net_by_c.items(), key=lambda x: -x[1]):
        va.append([f"Kitted AMPs - {cname}", net / 2,
                   f"=SUMIFS({EN_NET},{EN_CON},\"{cname}\")/2", f"=IF(B{r}=C{r},\"OK\",\"CHECK\")"])
        r += 1
    va.append(["Default selected week", "2026-W29"])
    va.append(["Default selected quarter", "2026-Q3"])
    # fix selector defaults: Dashboard C5/C6 reference Validation B{r} rows
    db["C5"] = f"=Validation!B{r}"
    db["C6"] = f"=Validation!B{r + 1}"
    for rowv in va.iter_rows(min_row=2):
        for c in rowv:
            c.font = AR
    va.column_dimensions["A"].width = 36
    va.column_dimensions["B"].width = 26
    va.column_dimensions["C"].width = 18

    wb.save(out_path)
    print(f"saved {out_path}: COOIS rows={n_coois}, engine capacity={COOIS_CAP}, "
          f"map rows={MAP_CAP}, contracts={N_CONTRACTS}, weeks={len(wks)}, quarters={len(qs)}")
    print("expected totals:", dict(net_by_c), "-> AMPs:", total_amps)


if __name__ == "__main__":
    build(sys.argv[1], sys.argv[2], sys.argv[3])
