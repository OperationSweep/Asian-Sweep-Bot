"""Discover the workbook's layout, then normalise it into Job objects.

The layout is discovered rather than assumed, so a sheet with extra rows above
the header, or a different number of fibre pairs, still parses. What is assumed
is the *shape*: a label column on the left, one column per work order, and one
row per component material.

Discovered from the sample workbook (ZPRO.xlsx):

    row 1  GFF Allocation   M1/M2   M3/M4   ...  M47/M48    <- fibre pairs
    row 2  Work Order s/n   <order> <order> ...             <- work orders
    row 3  Amp Tray s/n     <serial> <serial> ...           <- component rows
    ...                                                        (col A = material)
    row 37 <material code>  <serial> <serial> ...
    row 38                  1  2  3  ... 24                 <- index row, ignored
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import List, Optional, Sequence, Tuple

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ..models import (
    Anomaly,
    ComponentSerial,
    Job,
    KittingError,
    Severity,
    Workbook,
    normalise_serial,
    normalise_work_order,
)

log = logging.getLogger(__name__)

DEFAULT_WORK_ORDER_LABEL = "work order"


def _find_work_order_row(sheet: Worksheet, label: str) -> int:
    """Find the row whose first cell labels the work-order line."""
    needle = label.strip().lower()
    for row in range(1, min(sheet.max_row, 50) + 1):
        value = sheet.cell(row, 1).value
        if value and needle in str(value).strip().lower():
            return row
    raise KittingError(
        f"Could not find a row labelled like {label!r} in column A. "
        f"Set excel.work_order_label in config.yaml to the actual label."
    )


def _find_data_columns(sheet: Worksheet, work_order_row: int) -> List[int]:
    """Columns to the right of A that carry a work-order number."""
    columns = []
    for column in range(2, sheet.max_column + 1):
        value = sheet.cell(work_order_row, column).value
        if value is None or str(value).strip() == "":
            continue
        try:
            normalise_work_order(value)
        except ValueError:
            continue
        columns.append(column)
    if not columns:
        raise KittingError(
            f"Row {work_order_row} is labelled as the work-order row but holds "
            f"no order numbers."
        )
    return columns


def _no_columns_message(path: Path, sheet_title: str, work_order_row: int) -> str:
    """Explain an empty work-order row, which usually means an unresolved link.

    When the row is formulas, openpyxl reads only what Excel cached. A file
    saved without the linked workbook available caches nothing, and the row
    looks empty - which is a very different problem from a wrong label.
    """
    try:
        formulas = openpyxl.load_workbook(path, data_only=False)
        sheet = formulas[sheet_title]
    except Exception:
        return ""

    computed = [
        column for column in range(2, sheet.max_column + 1)
        if isinstance(sheet.cell(work_order_row, column).value, str)
        and str(sheet.cell(work_order_row, column).value).startswith("=")
    ]
    if not computed:
        return ""

    targets = _external_link_targets(formulas)
    source = f" ({targets[0]})" if targets else ""
    return (
        f" Row {work_order_row} holds {len(computed)} formulas, so the numbers "
        f"come from another workbook{source} and none of them has a cached "
        f"value here. Open this file in Excel with that source reachable, let "
        f"the links refresh, save, and run again."
    )


def _find_component_rows(sheet: Worksheet, work_order_row: int) -> List[int]:
    """Rows below the work-order row that name a material in column A.

    Stops at the first gap in column A, which is what separates the component
    block from the trailing index row and any notes below it.
    """
    rows = []
    for row in range(work_order_row + 1, sheet.max_row + 1):
        value = sheet.cell(row, 1).value
        if value is None or str(value).strip() == "":
            break
        rows.append(row)
    if not rows:
        raise KittingError("No component rows found below the work-order row.")
    return rows


def _find_stray_cells(
    sheet: Worksheet, last_component_row: int, columns: List[int]
) -> List[Tuple[str, int, object]]:
    """Any populated cell below the component block.

    The sample workbook has one: N45 holds a lone serial seven rows under the
    data. Values like that are never silently swept into a posting - they are
    surfaced so a human decides what they were meant to be.
    """
    stray = []
    for row in range(last_component_row + 1, sheet.max_row + 1):
        for column in range(1, sheet.max_column + 1):
            value = sheet.cell(row, column).value
            if value is None or str(value).strip() == "":
                continue
            # The index row (1, 2, 3, ... under each work order) is furniture.
            if column in columns and isinstance(value, (int, float)):
                continue
            stray.append((sheet.cell(row, column).coordinate, column, value))
    return stray



def _external_link_targets(book) -> List[str]:
    """Workbooks this one pulls values from, if any."""
    targets = []
    for link in getattr(book, "_external_links", None) or []:
        target = getattr(getattr(link, "file_link", None), "Target", None)
        if target:
            targets.append(str(target))
    return targets


def _check_volatile_sources(
    path: Path, sheet_title: str, work_order_row: int, columns: List[int]
) -> List[Anomaly]:
    """Warn when the work-order numbers are formulas rather than typed values.

    In the sample workbook rows 1-3 are formulas, and the work-order row is a
    link to another workbook. openpyxl can only read the value Excel cached the
    last time the file was opened with that source reachable. A stale cache
    would post serials against the wrong order numbers, so this is surfaced
    rather than trusted silently.
    """
    anomalies: List[Anomaly] = []
    try:
        formulas = openpyxl.load_workbook(path, data_only=False)
    except Exception:
        return anomalies

    if sheet_title not in formulas.sheetnames:
        return anomalies
    sheet = formulas[sheet_title]

    computed = [
        sheet.cell(work_order_row, column).coordinate
        for column in columns
        if isinstance(sheet.cell(work_order_row, column).value, str)
        and str(sheet.cell(work_order_row, column).value).startswith("=")
    ]
    if not computed:
        return anomalies

    targets = _external_link_targets(formulas)
    source = f" from {targets[0]}" if targets else ""
    anomalies.append(Anomaly(
        Severity.WARNING, "COMPUTED_WORK_ORDERS",
        f"{len(computed)} work order numbers are formulas{source}, not typed "
        f"values. The numbers read here are whatever Excel cached last - open "
        f"the workbook in Excel, let the link refresh, and save it before "
        f"running, or the batch may post against stale order numbers.",
        cell=computed[0],
    ))
    return anomalies


def read_workbook(
    path: str | Path,
    sheet_name: Optional[str] = None,
    work_order_label: str = DEFAULT_WORK_ORDER_LABEL,
    ignore_cells: Sequence[str] = (),
) -> Workbook:
    """Parse the workbook into jobs, recording anomalies rather than guessing.

    ignore_cells names cells already identified as leftovers in the template -
    they are still detected and logged, just not raised as warnings every run.
    """
    path = Path(path)
    if not path.exists():
        raise KittingError(f"spreadsheet not found: {path}")

    book = openpyxl.load_workbook(path, data_only=True)
    if sheet_name and sheet_name not in book.sheetnames:
        raise KittingError(
            f"worksheet {sheet_name!r} not in {path.name}. "
            f"Available: {book.sheetnames}"
        )
    sheet = book[sheet_name] if sheet_name else book.worksheets[0]

    work_order_row = _find_work_order_row(sheet, work_order_label)
    fibre_pair_row = work_order_row - 1 if work_order_row > 1 else None
    try:
        columns = _find_data_columns(sheet, work_order_row)
    except KittingError as exc:
        raise KittingError(
            str(exc) + _no_columns_message(path, sheet.title, work_order_row)
        ) from exc
    component_rows = _find_component_rows(sheet, work_order_row)

    result = Workbook(path=str(path), sheet_name=sheet.title)
    result.anomalies.extend(
        _check_volatile_sources(path, sheet.title, work_order_row, columns)
    )

    # Which materials carry a serial anywhere? A material that carries none is
    # non-serialised in this build, and is the expected cause of SAP's
    # "Serial Number missing for <material>" message.
    for row in component_rows:
        material = str(sheet.cell(row, 1).value).strip()
        has_serial = any(
            normalise_serial(sheet.cell(row, column).value) is not None
            for column in columns
        )
        target = (
            result.serialised_materials if has_serial
            else result.non_serialised_materials
        )
        if material not in target:
            target.append(material)

    seen_work_orders: dict[str, str] = {}

    for index, column in enumerate(columns, start=1):
        letter = get_column_letter(column)
        raw_order = sheet.cell(work_order_row, column).value
        work_order = normalise_work_order(raw_order)

        fibre_pair = ""
        if fibre_pair_row:
            label = sheet.cell(fibre_pair_row, column).value
            fibre_pair = str(label).strip() if label else ""

        job = Job(
            work_order=work_order,
            fibre_pair=fibre_pair,
            column_letter=letter,
            column_index=index,
        )

        if work_order in seen_work_orders:
            job.anomalies.append(Anomaly(
                Severity.BLOCKER, "DUPLICATE_WORK_ORDER",
                f"work order {work_order} also appears in column "
                f"{seen_work_orders[work_order]}",
                cell=f"{letter}{work_order_row}", work_order=work_order,
            ))
        seen_work_orders[work_order] = letter

        sequence = 0
        for row in component_rows:
            material = str(sheet.cell(row, 1).value).strip()
            cell = sheet.cell(row, column)
            serial = normalise_serial(cell.value)

            if serial is None:
                # Only a gap worth reporting if this material is serialised
                # elsewhere in the sheet - otherwise it is simply a bulk part.
                if material in result.serialised_materials:
                    raw = cell.value
                    job.anomalies.append(Anomaly(
                        Severity.WARNING, "MISSING_SERIAL",
                        f"{material} has no serial for {work_order} "
                        f"(cell holds {raw!r}), but is serialised on other "
                        f"orders - it will not be sent to SAP",
                        cell=cell.coordinate, work_order=work_order,
                    ))
                continue

            sequence += 1
            job.serials.append(ComponentSerial(
                material=material, serial=serial,
                sheet_row=row, sequence=sequence,
            ))

        if not job.serials:
            job.anomalies.append(Anomaly(
                Severity.BLOCKER, "NO_SERIALS",
                f"work order {work_order} has no serial numbers at all",
                cell=f"{letter}{work_order_row}", work_order=work_order,
            ))

        result.jobs.append(job)

    ignored = {c.strip().upper() for c in ignore_cells}
    for coordinate, column, value in _find_stray_cells(
        sheet, component_rows[-1], columns
    ):
        if coordinate.upper() in ignored:
            log.info(
                "%s holds %r below the data block; ignored by configuration",
                coordinate, value,
            )
            continue
        owner = ""
        if column in columns:
            owning = normalise_work_order(sheet.cell(work_order_row, column).value)
            owner = f" under work order {owning}"
        result.anomalies.append(Anomaly(
            Severity.WARNING, "STRAY_CELL",
            f"{value!r} sits below the component block{owner}. It is not part "
            f"of any work order and will be ignored - confirm it is not a "
            f"serial that belongs in the grid.",
            cell=coordinate,
        ))

    return result
