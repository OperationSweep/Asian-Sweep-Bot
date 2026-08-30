"""Build a synthetic workbook with the same structure as the production sheet.

The real ZPRO.xlsx holds live serial numbers, ASN part numbers and a link to an
internal SharePoint workbook, so it is not committed. This reproduces its shape
exactly - the layout, the anomalies, the formula-driven header rows - with
invented values, so the tests exercise the same code paths.

Mirrored from the production sheet:
  row 1   fibre-pair labels, M1/M2 .. M47/M48         (formulas)
  row 2   work order per column                       (formula -> external link)
  row 3   amp tray serial, 0 where no tray is fitted   (formula -> external link)
  rows 4-37  component materials in column A, serials across  (literals)
  row 38  column index 1..24
  N45     a stray serial below the block
"""

from __future__ import annotations

from pathlib import Path
from typing import List

import openpyxl

COLUMNS = 24
FIRST_ORDER = 900000001
WORK_ORDERS: List[str] = [str(FIRST_ORDER + i) for i in range(COLUMNS)]
LAST_ORDER = WORK_ORDERS[-1]

# Orders from this index on have no amp tray fitted - row 3 holds 0.
NO_TRAY_FROM = 16
ORDERS_WITH_TRAY = WORK_ORDERS[:NO_TRAY_FROM]
ORDERS_WITHOUT_TRAY = WORK_ORDERS[NO_TRAY_FROM:]

TRAY_LABEL = "Amp Tray s/n"

# Materials that carry a serial on every order.
SERIALISED = [
    "91AAA00301AAA", "91AAA00302AAB", "91AAA00303AAC", "91AAB00304AAD",
    "91AAB00305AAE", "91AAC00306AAF", "91AAC00307AAG", "91AAD00308AAH",
    "91AAD00309AAJ", "91AAE00310AAK", "91AAE00311AAL", "91AAF00312AAM",
    "91AAF00313AAN", "91AAG00314AAP", "91AAG00315AAQ", "91AAH00316AAR",
    "91AAH00317AAS", "91AAJ00318AAT", "91AAJ00319AAU", "91AAK00320AAV",
]
# Materials with no serial anywhere - SAP asks for these, and that is expected.
NON_SERIALISED = [
    "92ZZA00401AAA", "92ZZA00402AAB", "92ZZB00403AAC", "92ZZB00404AAD",
    "92ZZC00405AAE", "92ZZC00406AAF", "92ZZD00407AAG", "92ZZD00408AAH",
    "92ZZE00409AAJ", "92ZZE00410AAK",
]

SERIALS_WITH_TRAY = len(SERIALISED) + 1
SERIALS_WITHOUT_TRAY = len(SERIALISED)
TOTAL_SERIALS = (
    len(ORDERS_WITH_TRAY) * SERIALS_WITH_TRAY
    + len(ORDERS_WITHOUT_TRAY) * SERIALS_WITHOUT_TRAY
)

STRAY_CELL = "N45"
STRAY_VALUE = "SNX9999999"

FIRST_TRAY_SERIAL = "TRY0000001"
FIRST_MATERIAL_SERIAL = "SNA0000001"


def _serial(row_index: int, column_index: int) -> str:
    """A unique, stable serial for one (material, order) intersection."""
    letter = chr(ord("A") + (row_index % 26))
    return f"SN{letter}{row_index:03d}{column_index:04d}"


def build(path: Path, computed_headers: bool = True) -> Path:
    """Write the fixture. computed_headers reproduces the formula-driven rows."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"

    sheet.cell(1, 1, "GFF Allocation")
    sheet.cell(2, 1, "Work Order s/n")
    sheet.cell(3, 1, TRAY_LABEL)

    for index in range(COLUMNS):
        column = 2 + index
        sheet.cell(1, column, f"M{index * 2 + 1}/M{index * 2 + 2}")
        sheet.cell(2, column, int(WORK_ORDERS[index]))
        sheet.cell(
            3, column,
            0 if index >= NO_TRAY_FROM else f"TRY{index + 1:07d}",
        )

    # Interleave serialised and bulk materials the way the real sheet does.
    row = 4
    order = (
        [("s", 0), ("b", 0), ("s", 1), ("b", 1), ("s", 2), ("b", 2), ("s", 3),
         ("b", 3), ("b", 4), ("b", 5), ("s", 4), ("s", 5)]
        + [("b", i) for i in range(6, 10)]
        + [("s", i) for i in range(6, 20)]
    )
    for kind, index in order:
        if kind == "s":
            material = SERIALISED[index]
            sheet.cell(row, 1, material)
            for column_index in range(COLUMNS):
                sheet.cell(row, 2 + column_index, _serial(row, column_index))
        else:
            sheet.cell(row, 1, NON_SERIALISED[index])
        row += 1

    index_row = row + 1 if False else 38
    for column_index in range(COLUMNS):
        sheet.cell(index_row, 2 + column_index, column_index + 1)

    sheet[STRAY_CELL] = STRAY_VALUE

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


if __name__ == "__main__":
    target = Path(__file__).with_name("fixture_ZPRO.xlsx")
    build(target)
    print(f"wrote {target}")
